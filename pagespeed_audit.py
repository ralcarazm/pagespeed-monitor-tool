#!/usr/bin/env python3

import argparse
import csv
import logging
import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

try:
    import tomllib
except ModuleNotFoundError:
    import tomli as tomllib


API_ENDPOINT = "https://www.googleapis.com/pagespeedonline/v5/runPagespeed"

FIELD_METRICS = {
    "LARGEST_CONTENTFUL_PAINT_MS": "lcp_ms",
    "CUMULATIVE_LAYOUT_SHIFT_SCORE": "cls",
    "INTERACTION_TO_NEXT_PAINT": "inp_ms",
    "FIRST_CONTENTFUL_PAINT_MS": "fcp_ms",
    "EXPERIMENTAL_TIME_TO_FIRST_BYTE": "ttfb_ms",
}

LAB_AUDITS = {
    "first-contentful-paint": "lab_fcp_ms",
    "largest-contentful-paint": "lab_lcp_ms",
    "speed-index": "lab_speed_index_ms",
    "total-blocking-time": "lab_tbt_ms",
    "cumulative-layout-shift": "lab_cls",
    "interactive": "lab_tti_ms",
}

VALID_STRATEGIES = {"mobile", "desktop"}

VALID_CATEGORIES = {
    "performance",
    "accessibility",
    "best-practices",
    "seo",
}

DEFAULT_RETRY_STATUS_CODES = {408, 429, 500, 502, 503, 504}

BASE_FIELD_ORDER = [
    "checked_at_utc",
    "status",
    "requested_input_url",
    "strategy",
    "error_type",
    "error_message",
    "psi_id",
    "analysis_utc_timestamp",
    "final_url",
    "lighthouse_version",
    "runtime_error_code",
    "runtime_error_message",
    "performance_score",
    "accessibility_score",
    "best_practices_score",
    "seo_score",
    "field_url_data_available",
    "field_url_data_status",
    "field_url_data_message",
    "field_origin_data_available",
    "field_origin_data_status",
    "field_origin_data_message",
]


def load_config(config_path: str) -> dict[str, Any]:
    path = Path(config_path)

    if not path.exists():
        raise FileNotFoundError(
            f"No se ha encontrado el fichero de configuración: {config_path}"
        )

    with path.open("rb") as file:
        return tomllib.load(file)


def setup_logging(config: dict[str, Any]) -> logging.Logger:
    logger = logging.getLogger("pagespeed_audit")
    logger.handlers.clear()
    logger.propagate = False

    log_enabled = bool(config.get("log_enabled", True))
    log_level_name = str(config.get("log_level", "INFO")).upper()
    log_level = getattr(logging, log_level_name, logging.INFO)

    logger.setLevel(log_level)

    formatter = logging.Formatter(
        "%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    console_handler = logging.StreamHandler()
    console_handler.setLevel(log_level)
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

    if log_enabled:
        log_dir = Path(config.get("log_dir", "logs"))
        log_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = log_dir / f"pagespeed_audit_{timestamp}.log"

        file_handler = logging.FileHandler(log_file, encoding="utf-8")
        file_handler.setLevel(log_level)
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)

        logger.info("Log creado en: %s", log_file)

    return logger


def validate_config(config: dict[str, Any]) -> None:
    urls = config.get("urls")

    if not isinstance(urls, list) or not urls:
        raise ValueError("El fichero de configuración debe incluir una lista no vacía de URLs.")

    for url in urls:
        if not isinstance(url, str) or not url.startswith(("http://", "https://")):
            raise ValueError(f"URL no válida: {url}")

    strategies = config.get("strategies", ["mobile"])

    if not isinstance(strategies, list) or not strategies:
        raise ValueError("'strategies' debe ser una lista no vacía.")

    invalid_strategies = set(strategies) - VALID_STRATEGIES

    if invalid_strategies:
        raise ValueError(
            f"Estrategias no válidas: {', '.join(sorted(invalid_strategies))}"
        )

    categories = config.get("categories", ["performance"])

    if not isinstance(categories, list) or not categories:
        raise ValueError("'categories' debe ser una lista no vacía.")

    invalid_categories = set(categories) - VALID_CATEGORIES

    if invalid_categories:
        raise ValueError(
            f"Categorías no válidas: {', '.join(sorted(invalid_categories))}"
        )

    numeric_checks = {
        "sleep": (float, 1.0, 0, True),
        "timeout": (int, 120, 0, False),
        "retries": (int, 3, 0, True),
        "retry_backoff": (float, 2.0, 0, True),
    }

    for key, caster_data in numeric_checks.items():
        caster, default, minimum, allow_equal = caster_data

        try:
            value = caster(config.get(key, default))
        except (TypeError, ValueError) as exc:
            raise ValueError(f"'{key}' debe ser numérico.") from exc

        if allow_equal and value < minimum:
            raise ValueError(f"'{key}' no puede ser menor que {minimum}.")

        if not allow_equal and value <= minimum:
            raise ValueError(f"'{key}' debe ser mayor que {minimum}.")

    retry_status_codes = config.get(
        "retry_status_codes",
        sorted(DEFAULT_RETRY_STATUS_CODES),
    )

    if not isinstance(retry_status_codes, list):
        raise ValueError("'retry_status_codes' debe ser una lista de códigos HTTP.")

    for code in retry_status_codes:
        if not isinstance(code, int):
            raise ValueError("'retry_status_codes' debe contener solo números enteros.")

    csv_delimiter = config.get("csv_delimiter", ";")

    if not isinstance(csv_delimiter, str) or len(csv_delimiter) != 1:
        raise ValueError("'csv_delimiter' debe ser un único carácter.")

    csv_encoding = config.get("csv_encoding", "utf-8-sig")

    if not isinstance(csv_encoding, str) or not csv_encoding.strip():
        raise ValueError("'csv_encoding' debe ser una cadena de texto no vacía.")


def safe_get(data: dict[str, Any], *path: str) -> Any:
    current: Any = data

    for key in path:
        if not isinstance(current, dict):
            return None

        current = current.get(key)

    return current


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def clean_text_value(value: str) -> str:
    return value.replace("\u00a0", " ")


def clean_excel_value(value: Any) -> Any:
    if value is None:
        return ""

    if isinstance(value, str):
        return clean_text_value(value)

    return value


def clean_csv_value(value: Any, decimal_comma: bool = True) -> Any:
    if value is None:
        return ""

    if isinstance(value, bool):
        return str(value).lower()

    if isinstance(value, str):
        return clean_text_value(value)

    if decimal_comma and isinstance(value, float):
        return f"{value:.15g}".replace(".", ",")

    return value


def normalise_field_value(metric_key: str, value: Any) -> Any:
    if value is None:
        return None

    if metric_key == "CUMULATIVE_LAYOUT_SHIFT_SCORE":
        return value / 100

    return value


def add_seconds_value(result: dict[str, Any], ms_column_name: str, value: Any) -> None:
    if not ms_column_name.endswith("_ms"):
        return

    seconds_column_name = ms_column_name[:-3] + "_s"

    if is_number(value):
        result[seconds_column_name] = value / 1000
    else:
        result[seconds_column_name] = None


def build_lab_field_order(include_display_values: bool) -> list[str]:
    fieldnames: list[str] = []

    for column_name in LAB_AUDITS.values():
        fieldnames.append(column_name)

        if column_name.endswith("_ms"):
            fieldnames.append(column_name[:-3] + "_s")

        if include_display_values:
            fieldnames.append(f"{column_name}_display")

        fieldnames.append(f"{column_name}_score")

    return fieldnames


def build_field_metric_order(prefix: str) -> list[str]:
    fieldnames = [
        f"{prefix}_id",
        f"{prefix}_overall_category",
        f"{prefix}_data_available",
        f"{prefix}_data_status",
        f"{prefix}_data_message",
    ]

    for short_name in FIELD_METRICS.values():
        fieldnames.append(f"{prefix}_{short_name}")
        fieldnames.append(f"{prefix}_{short_name}_status")

        if short_name.endswith("_ms"):
            seconds_name = short_name[:-3] + "_s"
            fieldnames.append(f"{prefix}_{seconds_name}")
            fieldnames.append(f"{prefix}_{seconds_name}_status")

        fieldnames.append(f"{prefix}_{short_name}_category")
        fieldnames.append(f"{prefix}_{short_name}_good_pct")
        fieldnames.append(f"{prefix}_{short_name}_needs_improvement_pct")
        fieldnames.append(f"{prefix}_{short_name}_poor_pct")

    return fieldnames


def build_full_field_order(include_display_values: bool) -> list[str]:
    fieldnames = list(BASE_FIELD_ORDER)

    for field in build_lab_field_order(include_display_values):
        if field not in fieldnames:
            fieldnames.append(field)

    for prefix in ["field_url", "field_origin"]:
        for field in build_field_metric_order(prefix):
            if field not in fieldnames:
                fieldnames.append(field)

    return fieldnames


def build_request_params(
    url: str,
    strategy: str,
    api_key: str | None,
    categories: list[str],
    locale: str,
) -> list[tuple[str, str]]:
    params: list[tuple[str, str]] = [
        ("url", url),
        ("strategy", strategy),
        ("locale", locale),
    ]

    for category in categories:
        params.append(("category", category))

    if api_key:
        params.append(("key", api_key))

    return params


def get_response_error_payload(response: requests.Response) -> Any:
    try:
        return response.json()
    except ValueError:
        return response.text


def format_http_error(
    response: requests.Response,
    url: str,
    strategy: str,
) -> str:
    error_payload = get_response_error_payload(response)

    return (
        f"Error HTTP {response.status_code} al analizar {url} "
        f"({strategy}): {error_payload}"
    )


def sleep_before_retry(attempt: int, retry_backoff: float) -> None:
    delay = retry_backoff * (2 ** (attempt - 1))

    if delay > 0:
        time.sleep(delay)


def fetch_pagespeed(
    session: requests.Session,
    url: str,
    strategy: str,
    api_key: str | None,
    categories: list[str],
    locale: str,
    timeout: int,
    retries: int,
    retry_backoff: float,
    retry_status_codes: set[int],
    logger: logging.Logger,
) -> dict[str, Any]:
    params = build_request_params(
        url=url,
        strategy=strategy,
        api_key=api_key,
        categories=categories,
        locale=locale,
    )

    max_attempts = retries + 1

    for attempt in range(1, max_attempts + 1):
        try:
            logger.info(
                "Petición API intento %s/%s | URL=%s | strategy=%s",
                attempt,
                max_attempts,
                url,
                strategy,
            )

            response = session.get(
                API_ENDPOINT,
                params=params,
                timeout=timeout,
            )

            if (
                response.status_code in retry_status_codes
                and attempt < max_attempts
            ):
                logger.warning(
                    "Error temporal HTTP %s | URL=%s | strategy=%s | reintento=%s/%s",
                    response.status_code,
                    url,
                    strategy,
                    attempt,
                    retries,
                )
                sleep_before_retry(attempt, retry_backoff)
                continue

            try:
                response.raise_for_status()
            except requests.HTTPError as exc:
                raise RuntimeError(
                    format_http_error(response, url, strategy)
                ) from exc

            try:
                return response.json()
            except ValueError as exc:
                raise RuntimeError(
                    f"La API no ha devuelto JSON válido para {url} ({strategy})."
                ) from exc

        except (requests.Timeout, requests.ConnectionError) as exc:
            if attempt < max_attempts:
                logger.warning(
                    "Error temporal de red | URL=%s | strategy=%s | reintento=%s/%s | detalle=%s",
                    url,
                    strategy,
                    attempt,
                    retries,
                    exc,
                )
                sleep_before_retry(attempt, retry_backoff)
                continue

            raise RuntimeError(
                f"Error de red al analizar {url} ({strategy}) "
                f"tras {max_attempts} intento(s): {exc}"
            ) from exc

        except requests.RequestException as exc:
            raise RuntimeError(
                f"Error de petición al analizar {url} ({strategy}): {exc}"
            ) from exc

    raise RuntimeError(
        f"No se ha podido analizar {url} ({strategy}) tras {max_attempts} intento(s)."
    )


def extract_category_scores(data: dict[str, Any]) -> dict[str, Any]:
    categories = safe_get(data, "lighthouseResult", "categories") or {}

    result: dict[str, Any] = {}

    for category_id in ["performance", "accessibility", "best-practices", "seo"]:
        score = safe_get(categories, category_id, "score")

        result[f"{category_id.replace('-', '_')}_score"] = (
            round(score * 100, 2) if is_number(score) else None
        )

    return result


def extract_lab_audits(
    data: dict[str, Any],
    include_display_values: bool = False,
) -> dict[str, Any]:
    audits = safe_get(data, "lighthouseResult", "audits") or {}

    result: dict[str, Any] = {}

    for audit_id, column_name in LAB_AUDITS.items():
        audit = audits.get(audit_id, {})
        numeric_value = audit.get("numericValue")

        result[column_name] = numeric_value

        if column_name.endswith("_ms"):
            add_seconds_value(result, column_name, numeric_value)

        if include_display_values:
            result[f"{column_name}_display"] = audit.get("displayValue")

        result[f"{column_name}_score"] = audit.get("score")

    return result


def get_field_data_message(
    block_name: str,
    block: dict[str, Any],
    metrics: dict[str, Any],
) -> str:
    if metrics:
        return "Datos de campo disponibles."

    if not block:
        if block_name == "loadingExperience":
            return "Datos de campo no disponibles para la URL analizada."
        return "Datos de campo no disponibles para el origen analizado."

    overall_category = block.get("overall_category")

    if overall_category == "NONE":
        return "CrUX no dispone de muestra suficiente para calcular métricas de campo."

    return "La respuesta de PageSpeed Insights no incluye métricas de campo."


def extract_field_metrics(
    data: dict[str, Any],
    block_name: str,
    prefix: str,
    logger: logging.Logger | None = None,
    requested_url: str | None = None,
    strategy: str | None = None,
) -> dict[str, Any]:
    block = data.get(block_name) or {}
    metrics = block.get("metrics") or {}

    data_available = bool(metrics)
    data_status = "available" if data_available else "not_available"
    data_message = get_field_data_message(block_name, block, metrics)

    if logger and not data_available:
        logger.info(
            "Datos de campo no disponibles | block=%s | URL=%s | strategy=%s | motivo=%s",
            block_name,
            requested_url,
            strategy,
            data_message,
        )

    result: dict[str, Any] = {
        f"{prefix}_id": block.get("id"),
        f"{prefix}_overall_category": block.get("overall_category"),
        f"{prefix}_data_available": data_available,
        f"{prefix}_data_status": data_status,
        f"{prefix}_data_message": data_message,
    }

    for metric_key, short_name in FIELD_METRICS.items():
        metric = metrics.get(metric_key)

        if not metric:
            result[f"{prefix}_{short_name}"] = None
            result[f"{prefix}_{short_name}_status"] = "not_available"
            result[f"{prefix}_{short_name}_category"] = "not_available"
            result[f"{prefix}_{short_name}_good_pct"] = None
            result[f"{prefix}_{short_name}_needs_improvement_pct"] = None
            result[f"{prefix}_{short_name}_poor_pct"] = None

            if short_name.endswith("_ms"):
                seconds_name = short_name[:-3] + "_s"
                result[f"{prefix}_{seconds_name}"] = None
                result[f"{prefix}_{seconds_name}_status"] = "not_available"

            continue

        percentile = metric.get("percentile")
        value = normalise_field_value(metric_key, percentile)
        distributions = metric.get("distributions") or []

        result[f"{prefix}_{short_name}"] = value
        result[f"{prefix}_{short_name}_status"] = "available"

        if short_name.endswith("_ms"):
            seconds_name = short_name[:-3] + "_s"
            result[f"{prefix}_{seconds_name}"] = (
                value / 1000 if is_number(value) else None
            )
            result[f"{prefix}_{seconds_name}_status"] = "available"

        result[f"{prefix}_{short_name}_category"] = (
            metric.get("category") or "not_available"
        )

        result[f"{prefix}_{short_name}_good_pct"] = (
            round(distributions[0].get("proportion", 0) * 100, 2)
            if len(distributions) > 0
            else None
        )

        result[f"{prefix}_{short_name}_needs_improvement_pct"] = (
            round(distributions[1].get("proportion", 0) * 100, 2)
            if len(distributions) > 1
            else None
        )

        result[f"{prefix}_{short_name}_poor_pct"] = (
            round(distributions[2].get("proportion", 0) * 100, 2)
            if len(distributions) > 2
            else None
        )

    return result


def extract_result(
    data: dict[str, Any],
    requested_url: str,
    strategy: str,
    include_display_values: bool,
    logger: logging.Logger,
) -> dict[str, Any]:
    runtime_error_code = safe_get(data, "lighthouseResult", "runtimeError", "code")
    runtime_error_message = safe_get(data, "lighthouseResult", "runtimeError", "message")

    row: dict[str, Any] = {
        "checked_at_utc": datetime.now(timezone.utc).isoformat(),
        "status": "ok",
        "requested_input_url": requested_url,
        "strategy": strategy,
        "error_type": "sin_error",
        "error_message": "sin_error",
        "psi_id": data.get("id"),
        "analysis_utc_timestamp": data.get("analysisUTCTimestamp"),
        "final_url": safe_get(data, "lighthouseResult", "finalUrl"),
        "lighthouse_version": safe_get(data, "lighthouseResult", "lighthouseVersion"),
        "runtime_error_code": runtime_error_code or "sin_error",
        "runtime_error_message": runtime_error_message or "sin_error",
    }

    row.update(extract_category_scores(data))

    row.update(
        extract_lab_audits(
            data,
            include_display_values=include_display_values,
        )
    )

    row.update(
        extract_field_metrics(
            data,
            "loadingExperience",
            "field_url",
            logger=logger,
            requested_url=requested_url,
            strategy=strategy,
        )
    )

    row.update(
        extract_field_metrics(
            data,
            "originLoadingExperience",
            "field_origin",
            logger=logger,
            requested_url=requested_url,
            strategy=strategy,
        )
    )

    return row


def build_error_row(
    requested_url: str,
    strategy: str,
    error: Exception,
) -> dict[str, Any]:
    return {
        "checked_at_utc": datetime.now(timezone.utc).isoformat(),
        "status": "error",
        "requested_input_url": requested_url,
        "strategy": strategy,
        "error_type": type(error).__name__,
        "error_message": str(error),
        "runtime_error_code": "no_ejecutado",
        "runtime_error_message": "El análisis no se ha completado correctamente.",
        "field_url_data_available": False,
        "field_url_data_status": "not_evaluated",
        "field_url_data_message": "No se han recuperado datos de campo porque la petición ha fallado.",
        "field_origin_data_available": False,
        "field_origin_data_status": "not_evaluated",
        "field_origin_data_message": "No se han recuperado datos de campo porque la petición ha fallado.",
    }


def fill_control_value(field_name: str, value: Any) -> Any:
    if value not in (None, ""):
        return value

    if field_name.endswith("_status"):
        return "not_available"

    if field_name.endswith("_category"):
        return "not_available"

    if field_name.endswith("_message"):
        return "not_available"

    if field_name in {"error_type", "error_message"}:
        return "sin_error"

    if field_name in {"runtime_error_code", "runtime_error_message"}:
        return "sin_error"

    return value


def prepare_row_for_output(
    row: dict[str, Any],
    fieldnames: list[str],
    fill_control_values: bool,
) -> dict[str, Any]:
    prepared = {}

    for field_name in fieldnames:
        value = row.get(field_name)

        if fill_control_values:
            value = fill_control_value(field_name, value)

        prepared[field_name] = value

    return prepared


def build_final_fieldnames(
    rows: list[dict[str, Any]],
    full_fieldnames: list[str],
    drop_empty_columns: bool = False,
) -> list[str]:
    fieldnames = list(full_fieldnames)

    for row in rows:
        for key in row.keys():
            if key not in fieldnames:
                fieldnames.append(key)

    if not drop_empty_columns:
        return fieldnames

    filtered_fieldnames: list[str] = []

    for field in fieldnames:
        if field in BASE_FIELD_ORDER:
            filtered_fieldnames.append(field)
            continue

        has_value = any(
            row.get(field) not in (None, "")
            for row in rows
        )

        if has_value:
            filtered_fieldnames.append(field)

    return filtered_fieldnames


def initialise_csv(
    output_path: str,
    fieldnames: list[str],
    delimiter: str,
    encoding: str,
) -> None:
    output = Path(output_path)

    if output.parent:
        output.parent.mkdir(parents=True, exist_ok=True)

    with output.open("w", encoding=encoding, newline="") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=fieldnames,
            delimiter=delimiter,
            extrasaction="ignore",
        )
        writer.writeheader()


def append_csv_row(
    row: dict[str, Any],
    output_path: str,
    fieldnames: list[str],
    delimiter: str,
    encoding: str,
    decimal_comma: bool,
    fill_control_values: bool,
) -> None:
    prepared_row = prepare_row_for_output(
        row=row,
        fieldnames=fieldnames,
        fill_control_values=fill_control_values,
    )

    cleaned_row = {
        key: clean_csv_value(value, decimal_comma=decimal_comma)
        for key, value in prepared_row.items()
    }

    with Path(output_path).open("a", encoding=encoding, newline="") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=fieldnames,
            delimiter=delimiter,
            extrasaction="ignore",
        )
        writer.writerow(cleaned_row)


def write_csv_full(
    rows: list[dict[str, Any]],
    output_path: str,
    fieldnames: list[str],
    delimiter: str,
    encoding: str,
    decimal_comma: bool,
    fill_control_values: bool,
) -> None:
    if not rows:
        return

    output = Path(output_path)

    if output.parent:
        output.parent.mkdir(parents=True, exist_ok=True)

    prepared_rows = [
        prepare_row_for_output(
            row=row,
            fieldnames=fieldnames,
            fill_control_values=fill_control_values,
        )
        for row in rows
    ]

    cleaned_rows = [
        {
            key: clean_csv_value(value, decimal_comma=decimal_comma)
            for key, value in row.items()
        }
        for row in prepared_rows
    ]

    with output.open("w", encoding=encoding, newline="") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=fieldnames,
            delimiter=delimiter,
            extrasaction="ignore",
        )
        writer.writeheader()
        writer.writerows(cleaned_rows)


def get_excel_number_format(field_name: str) -> str | None:
    if field_name.endswith("_ms"):
        return "0.000"

    if field_name.endswith("_s"):
        return "0.000"

    if field_name.endswith("_pct"):
        return "0.00"

    if field_name.endswith("_score"):
        return "0.00"

    if field_name.endswith("_cls"):
        return "0.000"

    return None


def write_xlsx(
    rows: list[dict[str, Any]],
    output_path: str,
    fieldnames: list[str],
    fill_control_values: bool = True,
) -> None:
    if not rows:
        return

    output = Path(output_path)

    if output.parent:
        output.parent.mkdir(parents=True, exist_ok=True)

    prepared_rows = [
        prepare_row_for_output(
            row=row,
            fieldnames=fieldnames,
            fill_control_values=fill_control_values,
        )
        for row in rows
    ]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PageSpeed Results"

    worksheet.append(fieldnames)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in prepared_rows:
        worksheet.append(
            [
                clean_excel_value(row.get(field_name))
                for field_name in fieldnames
            ]
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_index, field_name in enumerate(fieldnames, start=1):
        column_letter = get_column_letter(column_index)
        number_format = get_excel_number_format(field_name)

        if number_format:
            for row_index in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_index, column=column_index)

                if is_number(cell.value):
                    cell.number_format = number_format

        max_length = len(field_name)

        for row_index in range(2, worksheet.max_row + 1):
            value = worksheet.cell(row=row_index, column=column_index).value

            if value is not None:
                max_length = max(max_length, len(str(value)))

        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12),
            60,
        )

    workbook.save(output)


def write_final_outputs(
    rows: list[dict[str, Any]],
    full_fieldnames: list[str],
    csv_output: str,
    xlsx_output: str,
    write_csv_file: bool,
    write_xlsx_file: bool,
    csv_delimiter: str,
    csv_encoding: str,
    csv_decimal_comma: bool,
    fill_control_values: bool,
    drop_empty_columns: bool,
    logger: logging.Logger,
) -> None:
    final_fieldnames = build_final_fieldnames(
        rows=rows,
        full_fieldnames=full_fieldnames,
        drop_empty_columns=drop_empty_columns,
    )

    if write_csv_file:
        write_csv_full(
            rows=rows,
            output_path=csv_output,
            fieldnames=final_fieldnames,
            delimiter=csv_delimiter,
            encoding=csv_encoding,
            decimal_comma=csv_decimal_comma,
            fill_control_values=fill_control_values,
        )
        logger.info("CSV final generado: %s", csv_output)

    if write_xlsx_file:
        write_xlsx(
            rows=rows,
            output_path=xlsx_output,
            fieldnames=final_fieldnames,
            fill_control_values=fill_control_values,
        )
        logger.info("XLSX final generado: %s", xlsx_output)


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Audita URLs con Google PageSpeed Insights API "
            "usando un fichero de configuración TOML."
        )
    )

    parser.add_argument(
        "-c",
        "--config",
        default="config.toml",
        help="Ruta del fichero de configuración TOML. Por defecto: config.toml",
    )

    args = parser.parse_args()

    try:
        config = load_config(args.config)
        validate_config(config)
    except Exception as exc:
        print(f"Error en la configuración: {exc}", file=sys.stderr)
        return 1

    logger = setup_logging(config)

    logger.info("Inicio de ejecución.")
    logger.info("Fichero de configuración: %s", args.config)

    api_key = config.get("api_key") or os.getenv("PAGESPEED_API_KEY")
    urls = config["urls"]
    strategies = config.get("strategies", ["mobile"])
    categories = config.get("categories", ["performance"])
    locale = config.get("locale", "es")
    sleep_seconds = float(config.get("sleep", 1.0))
    timeout = int(config.get("timeout", 120))

    csv_output = config.get("output", "results/pagespeed_results.csv")
    xlsx_output = config.get(
        "xlsx_output",
        str(Path(csv_output).with_suffix(".xlsx")),
    )

    write_csv_file = bool(config.get("write_csv", True))
    write_xlsx_file = bool(config.get("write_xlsx", True))

    csv_delimiter = config.get("csv_delimiter", ";")
    csv_encoding = config.get("csv_encoding", "utf-8-sig")
    csv_decimal_comma = bool(config.get("csv_decimal_comma", True))

    include_display_values = bool(config.get("include_display_values", False))
    fill_control_values = bool(config.get("fill_control_values", True))
    drop_empty_columns = bool(config.get("drop_empty_columns", False))

    retries = int(config.get("retries", 3))
    retry_backoff = float(config.get("retry_backoff", 2.0))
    retry_status_codes = set(
        int(code)
        for code in config.get(
            "retry_status_codes",
            sorted(DEFAULT_RETRY_STATUS_CODES),
        )
    )

    full_fieldnames = build_full_field_order(
        include_display_values=include_display_values,
    )

    rows: list[dict[str, Any]] = []

    total = len(urls) * len(strategies)
    current = 0

    if write_csv_file:
        try:
            initialise_csv(
                output_path=csv_output,
                fieldnames=full_fieldnames,
                delimiter=csv_delimiter,
                encoding=csv_encoding,
            )
            logger.info("CSV progresivo inicializado: %s", csv_output)
        except Exception:
            logger.exception("No se ha podido inicializar el CSV progresivo.")
            return 1

    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": "pagespeed-audit/1.0",
        }
    )

    try:
        for url in urls:
            for strategy in strategies:
                current += 1

                logger.info(
                    "[%s/%s] Analizando URL=%s | strategy=%s",
                    current,
                    total,
                    url,
                    strategy,
                )

                try:
                    data = fetch_pagespeed(
                        session=session,
                        url=url,
                        strategy=strategy,
                        api_key=api_key,
                        categories=categories,
                        locale=locale,
                        timeout=timeout,
                        retries=retries,
                        retry_backoff=retry_backoff,
                        retry_status_codes=retry_status_codes,
                        logger=logger,
                    )

                    row = extract_result(
                        data=data,
                        requested_url=url,
                        strategy=strategy,
                        include_display_values=include_display_values,
                        logger=logger,
                    )

                    logger.info(
                        "Análisis correcto | URL=%s | strategy=%s",
                        url,
                        strategy,
                    )

                except Exception as exc:
                    logger.exception(
                        "Error analizando URL=%s | strategy=%s",
                        url,
                        strategy,
                    )

                    row = build_error_row(
                        requested_url=url,
                        strategy=strategy,
                        error=exc,
                    )

                rows.append(row)

                if write_csv_file:
                    try:
                        append_csv_row(
                            row=row,
                            output_path=csv_output,
                            fieldnames=full_fieldnames,
                            delimiter=csv_delimiter,
                            encoding=csv_encoding,
                            decimal_comma=csv_decimal_comma,
                            fill_control_values=fill_control_values,
                        )
                        logger.info("Fila añadida al CSV progresivo.")
                    except Exception:
                        logger.exception("Error al añadir fila al CSV progresivo.")
                        return 1

                time.sleep(sleep_seconds)

    except KeyboardInterrupt:
        logger.warning("Ejecución interrumpida por el usuario.")

        if rows:
            try:
                write_final_outputs(
                    rows=rows,
                    full_fieldnames=full_fieldnames,
                    csv_output=csv_output,
                    xlsx_output=xlsx_output,
                    write_csv_file=write_csv_file,
                    write_xlsx_file=write_xlsx_file,
                    csv_delimiter=csv_delimiter,
                    csv_encoding=csv_encoding,
                    csv_decimal_comma=csv_decimal_comma,
                    fill_control_values=fill_control_values,
                    drop_empty_columns=drop_empty_columns,
                    logger=logger,
                )

                logger.info("Resultados parciales guardados tras interrupción.")

            except Exception:
                logger.exception("No se han podido guardar los resultados parciales.")

        return 130

    finally:
        session.close()

    try:
        write_final_outputs(
            rows=rows,
            full_fieldnames=full_fieldnames,
            csv_output=csv_output,
            xlsx_output=xlsx_output,
            write_csv_file=write_csv_file,
            write_xlsx_file=write_xlsx_file,
            csv_delimiter=csv_delimiter,
            csv_encoding=csv_encoding,
            csv_decimal_comma=csv_decimal_comma,
            fill_control_values=fill_control_values,
            drop_empty_columns=drop_empty_columns,
            logger=logger,
        )
    except Exception:
        logger.exception("Error al generar los ficheros finales.")
        return 1

    logger.info("Ejecución finalizada correctamente.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())